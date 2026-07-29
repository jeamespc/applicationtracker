import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"

wb = openpyxl.Workbook()

# ---------------------------------------------------------------------------
# Data — transcribed from "The List.docx"
# ---------------------------------------------------------------------------
TIER_READY = "Ready to Apply"
TIER_APPLIED = "Applied"
TIER_STRONG = "Strong Fit"
TIER_DIFF = "Different Angle"
TIER_REJ_AUTO = "Rejected \u2014 Automatic (No Place)"
TIER_REJ_ASSESSED = "Rejected \u2014 Assessed"

rows = [
    # studio, status, tier, channel, contact, location, team_size, discipline, why_it_fits, portfolio_pieces
    ("Matteo Thun & Partners", "Not Started", TIER_READY, "Website", "", "Milan + Munich", "~70",
     "Architecture, interior, product, exhibition design",
     "Major pedigree (Thun co-founded Memphis Group with Sottsass). Live Junior Interior Designer posting right now, plus an actual connection through CrippaConcept.",
     ""),
    ("Formafantasma", "Not Started", TIER_READY, "Email", "work@formafantasma.com", "Milan/Rotterdam", "",
     "Design research (material & ecological)",
     "Research studio (Trimarchi & Farresin); treats design as material and ecological research rather than styling \u2014 form follows the research, not the other way around. Your strongest match.",
     "DEMY, Palazzo Bianco"),
    ("Atelierzero", "Not Started", TIER_READY, "Email", "info@atelierzero.it", "", "3 founders",
     "Residential / retail / product",
     "Built on \u201cemotional design with conceptual clarity.\u201d Language very close to your own design philosophy; also a registered internship host.",
     ""),
    ("Atelier Biagetti", "Not Started", TIER_READY, "Email", "info@atelierbiagetti.com", "Milan (Navigli)", "2 (duo)",
     "Furniture / art-object design",
     "Treat every project like a movie: surreal, theatrical, art-object furniture. Closest Milan equivalent to Supaform's small, authorial, provocative spirit.",
     "Supaform"),
    ("Migliore+Servetto", "Not Started", TIER_READY, "Email", "job@miglioreservetto.com", "", "",
     "Scenography / exhibition design",
     "Trained under Castiglioni; museums and \u201cnarrative spaces\u201d built around the visitor's journey. Probably your single best match \u2014 designing how people move through and encounter a space is their literal discipline.",
     ""),
    ("Park Associati", "Not Started", TIER_STRONG, "Website", "", "", "Mid-large",
     "Architecture & interior (corporate / retail)",
     "Corporate HQ and retail-scale architecture and interior design. Less personal-narrative fit, but size means real hiring turnover.",
     ""),
    ("Vudafieri-Saverino Partners", "Not Started", TIER_STRONG, "Email", "info@vudafierisaverino.it", "", "",
     "Retail, hospitality, restaurants",
     "Storytelling explicitly named as their method (ex-Sottsass pedigree). Their framing of \u201cConcept Design and Storytelling Strategy\u201d mirrors how you talk about your own process.",
     ""),
    ("Locatelli Partners", "Not Started", TIER_STRONG, "Email", "design@locatellipartners.com", "Milan + NY", "~60",
     "Architecture & interiors",
     "Material- and tech-research-driven. Prominent and well-resourced \u2014 a solid junior-role bet.",
     ""),
    ("Studiopepe", "Not Started", TIER_STRONG, "Email", "apply@studiopepe.info", "", "",
     "Interiors (retail / hospitality / residential)",
     "Eclectic, art-directed, materially rich interiors. Probably your closest single aesthetic match overall.",
     ""),
    ("Dimorestudio", "Not Started", TIER_STRONG, "Email", "careers@dimorestudio.eu", "", "40+",
     "Interiors (fashion / hospitality)",
     "Moody, atmospheric, \u201ctotal work of art\u201d interiors for major fashion and hospitality clients. Strong atmospheric match; has an active careers channel worth trying, though no live posting right now.",
     ""),
    ("Cristina Celestino", "Not Started", TIER_STRONG, "Email", "info@cristinacelestino.com", "", "Small (personal-name studio)",
     "Residential & product design",
     "Maximalist, color- and pattern-driven residential and product design. Closest to an authorial, small-scale practice like Supaform.",
     "Supaform"),
    ("Calvi Brambilla", "Not Started", TIER_STRONG, "Email", "studio@calvibrambilla.it", "", "Small team",
     "Interiors, exhibitions, product",
     "\u201cSubtle irony\u201d; teaching ties to Politecnico. Real exhibition-design chops at boutique scale \u2014 a good complement to Migliore+Servetto.",
     ""),
    ("Atomaa", "Not Started", TIER_STRONG, "Email", "jobs@atomaa.eu", "Milan + Edinburgh", "3 partners",
     "Residential / micro-living",
     "Hands-on material detailing. Boutique scale, close to your interior/spatial focus.",
     ""),
    ("Studio MI\u00b7LO", "Not Started", TIER_STRONG, "Email", "", "Milan + London", "Small boutique",
     "Interiors",
     "Evocative interiors. Good mentorship scale, though the least-researched name on this list \u2014 confirm details before applying.",
     ""),
    ("Lascia la Scia", "Not Started", TIER_STRONG, "Email", "", "", "5 founding women architects",
     "Architecture",
     "Runs its own coworking lab; warm, collaborative culture. Craft-forward, community-minded \u2014 feels like a place that mentors closely.",
     ""),
    ("Studio WOK", "Not Started", TIER_STRONG, "Email", "", "", "3 partners",
     "Architecture (habitat-focused)",
     "Narrative, habitat-focused architecture. Real, demonstrated investment in training junior designers \u2014 they head SPD's Interior Design Master.",
     ""),
    ("Stefano Boeri Architetti", "Not Started", TIER_STRONG, "Email", "hr@stefanoboeriarchitetti.net", "Milan/Shanghai/Tirana", "Large",
     "Ecological urbanism",
     "Large-scale ecological urbanism (Bosco Verticale). Fits your urban/spatial side better than interiors; has a jobs page and open channel.",
     ""),
    ("Studio Klass", "Not Started", TIER_STRONG, "Email", "careers@studioklass.com", "", "2 founders",
     "Product, interior, creative direction",
     "Teaching ties to IED. Smallest team here, with a direct pipeline from design education.",
     ""),
    ("CrippaConcept", "Not Started", TIER_DIFF, "Website", "", "", "",
     "Mobile-home / glamping manufacturer",
     "Not a design studio. Your existing warm connection via Prof. Trabattoni, thematically identical to Mechanisma, and the link into Matteo Thun & Partners.",
     "Mechanisma"),
    ("Studio Urquiola", "Not Started", TIER_DIFF, "Email", "recruiting@patriciaurquiola.com", "", "~40",
     "Architecture / interior / product",
     "Globally prominent (Patricia Urquiola); expressive, eclectic work. Big CV value; also worth trying for an actual junior role given its size.",
     ""),
    ("LLABB", "Not Started", TIER_DIFF, "Email", "internship@llabb.eu", "Genova (not Milan)", "",
     "Architecture (craft-rooted)",
     "Began as a carpentry workshop. Confirmed to host interns \u2014 worth the commute if hands-on material culture matters more than location.",
     ""),
    ("Onsitestudio", "Rejected", TIER_REJ_AUTO, "Email", "jobs@onsitestudio.it", "", "~18",
     "Architecture",
     "Material- and context-sensitive studio. Officially registered internship host with real open slots; required software matches your toolkit almost exactly \u2014 confirm why this was auto-rejected before fully closing the door.",
     ""),
]

STATUS_OPTIONS = ["Not Started", "Drafting", "Applied", "Responded", "Interview", "Offer", "Rejected", "Declined"]
TIER_OPTIONS = [TIER_READY, TIER_APPLIED, TIER_STRONG, TIER_DIFF, TIER_REJ_AUTO, TIER_REJ_ASSESSED]
CHANNEL_OPTIONS = ["Email", "Website"]

TIER_COLORS = {
    TIER_READY: "FFFF00",       # yellow — matches original doc
    TIER_APPLIED: "66CCCC",     # teal — distinct from Ready to Apply, matches HTML tracker
    TIER_STRONG: "FF9900",      # orange — matches original doc
    TIER_DIFF: "FFCC99",        # lighter orange — related but distinct from Strong Fit
    TIER_REJ_AUTO: "FF6B6B",    # red — matches original doc
    TIER_REJ_ASSESSED: "FF6B6B",
}
STATUS_COLORS = {
    "Not Started": None,
    "Drafting": "D9D9D9",
    "Applied": "FFF2A8",
    "Responded": "A9D18E",       # green — matches original doc's RESPONDED highlight
    "Interview": "9DC3E6",
    "Offer": "70AD47",
    "Rejected": "FF6B6B",
    "Declined": "808080",
}

HEADER_FILL = PatternFill("solid", fgColor="1F1F1F")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ---------------------------------------------------------------------------
# Sheet 1: Tracker
# ---------------------------------------------------------------------------
ws = wb.active
ws.title = "Tracker"

headers = ["Studio", "Status", "Fit Tier", "Channel", "Contact", "Location", "Team Size",
           "Discipline", "Why It Fits / Notes", "Portfolio Pieces to Feature",
           "Date Applied", "Follow-up Date", "Response Log"]
col_widths = [24, 13, 26, 10, 26, 16, 14, 26, 55, 20, 13, 13, 34]

for c, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(c)].width = w
ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"

for r, row in enumerate(rows, start=2):
    (studio, status, tier, channel, contact, location, team_size, discipline, why, pieces) = row
    values = [studio, status, tier, channel, contact, location, team_size, discipline, why, pieces, None, None, ""]
    for c, val in enumerate(values, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = Font(name=FONT_NAME, size=10)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = BORDER
    ws.cell(row=r, column=11).number_format = "yyyy-mm-dd"
    ws.cell(row=r, column=12).number_format = "yyyy-mm-dd"
    ws.row_dimensions[r].height = 46

last_row = len(rows) + 1
ws.auto_filter.ref = f"A1:M{last_row}"

# Data validation dropdowns
dv_status = DataValidation(type="list", formula1='"' + ",".join(STATUS_OPTIONS) + '"', allow_blank=True)
dv_tier = DataValidation(type="list", formula1='"' + ",".join(TIER_OPTIONS) + '"', allow_blank=True)
dv_channel = DataValidation(type="list", formula1='"' + ",".join(CHANNEL_OPTIONS) + '"', allow_blank=True)
ws.add_data_validation(dv_status)
ws.add_data_validation(dv_tier)
ws.add_data_validation(dv_channel)
dv_status.add(f"B2:B{last_row}")
dv_tier.add(f"C2:C{last_row}")
dv_channel.add(f"D2:D{last_row}")

# Conditional formatting: Fit Tier column (C)
for tier_val, color in TIER_COLORS.items():
    ws.conditional_formatting.add(
        f"C2:C{last_row}",
        CellIsRule(operator="equal", formula=[f'"{tier_val}"'], fill=PatternFill("solid", fgColor=color))
    )

# Conditional formatting: Status column (B)
for status_val, color in STATUS_COLORS.items():
    if color is None:
        continue
    ws.conditional_formatting.add(
        f"B2:B{last_row}",
        CellIsRule(operator="equal", formula=[f'"{status_val}"'], fill=PatternFill("solid", fgColor=color))
    )

# ---------------------------------------------------------------------------
# Sheet 2: Dashboard
# ---------------------------------------------------------------------------
ws2 = wb.create_sheet("Dashboard")
ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 10

title = ws2.cell(row=1, column=1, value="Application Tracker Dashboard")
title.font = Font(name=FONT_NAME, bold=True, size=14)
ws2.merge_cells("A1:B1")

ws2.cell(row=3, column=1, value="Total Studios").font = Font(name=FONT_NAME, bold=True, size=10)
ws2.cell(row=3, column=2, value=f"=COUNTA(Tracker!A2:A{last_row})").font = Font(name=FONT_NAME, size=10)

r = 5
ws2.cell(row=r, column=1, value="By Status").font = Font(name=FONT_NAME, bold=True, size=11)
r += 1
for status_val in STATUS_OPTIONS:
    ws2.cell(row=r, column=1, value=status_val).font = Font(name=FONT_NAME, size=10)
    cell = ws2.cell(row=r, column=2, value=f'=COUNTIF(Tracker!$B$2:$B${last_row},A{r})')
    cell.font = Font(name=FONT_NAME, size=10)
    color = STATUS_COLORS.get(status_val)
    if color:
        ws2.cell(row=r, column=1).fill = PatternFill("solid", fgColor=color)
    r += 1

r += 1
ws2.cell(row=r, column=1, value="By Fit Tier").font = Font(name=FONT_NAME, bold=True, size=11)
r += 1
for tier_val in TIER_OPTIONS:
    ws2.cell(row=r, column=1, value=tier_val).font = Font(name=FONT_NAME, size=10)
    cell = ws2.cell(row=r, column=2, value=f'=COUNTIF(Tracker!$C$2:$C${last_row},A{r})')
    cell.font = Font(name=FONT_NAME, size=10)
    ws2.cell(row=r, column=1).fill = PatternFill("solid", fgColor=TIER_COLORS[tier_val])
    r += 1

r += 1
ws2.cell(row=r, column=1, value="By Channel").font = Font(name=FONT_NAME, bold=True, size=11)
r += 1
for ch in CHANNEL_OPTIONS:
    ws2.cell(row=r, column=1, value=ch).font = Font(name=FONT_NAME, size=10)
    cell = ws2.cell(row=r, column=2, value=f'=COUNTIF(Tracker!$D$2:$D${last_row},A{r})')
    cell.font = Font(name=FONT_NAME, size=10)
    r += 1

r += 2
ws2.cell(row=r, column=1, value="Follow-ups Due (today or overdue)").font = Font(name=FONT_NAME, bold=True, size=11)
r += 1
ws2.cell(row=r, column=1, value="Count").font = Font(name=FONT_NAME, size=10)
fcell = ws2.cell(row=r, column=2,
                  value=f'=COUNTIFS(Tracker!$L$2:$L${last_row},"<="&TODAY(),Tracker!$L$2:$L${last_row},"<>")')
fcell.font = Font(name=FONT_NAME, size=10)

# ---------------------------------------------------------------------------
# Sheet 3: How To Use
# ---------------------------------------------------------------------------
ws3 = wb.create_sheet("How To Use")
ws3.column_dimensions["A"].width = 100
lines = [
    ("Application Tracker \u2014 How This Works", True, 14),
    ("", False, 10),
    ("This workbook rebuilds \u201cThe List\u201d as a working tracker instead of a static document.", False, 10),
    ("", False, 10),
    ("TABS", True, 12),
    ("Tracker \u2014 one row per studio. Edit Status, Date Applied, Follow-up Date, and Response Log as you go.", False, 10),
    ("Dashboard \u2014 live counts, pulled automatically from the Tracker tab. Nothing to edit here.", False, 10),
    ("", False, 10),
    ("COLUMNS TO EDIT AS YOU WORK", True, 12),
    ("Status \u2014 your day-to-day pipeline stage. Dropdown: " + ", ".join(STATUS_OPTIONS) + ".", False, 10),
    ("Date Applied / Follow-up Date \u2014 fill in when you send something; set Follow-up ~2 weeks out.", False, 10),
    ("Response Log \u2014 free text. Log what they said, interview dates, next steps.", False, 10),
    ("Portfolio Pieces to Feature \u2014 which of your projects (DEMY, Palazzo Bianco, Mechanisma, Supaform, etc.) to lead with for that studio.", False, 10),
    ("", False, 10),
    ("FIT TIER (carried over from The List, colour-coded same as the original)", True, 12),
    (TIER_READY + " (yellow) \u2014 top of the list: apply first.", False, 10),
    (TIER_APPLIED + " (teal) \u2014 already sent, waiting to hear back.", False, 10),
    (TIER_STRONG + " (orange) \u2014 strong match, not yet contacted.", False, 10),
    (TIER_DIFF + " (light orange) \u2014 adjacent / backup angle, not a core design-studio fit.", False, 10),
    (TIER_REJ_AUTO + " / " + TIER_REJ_ASSESSED + " (red) \u2014 not pursuing. Onsitestudio is pre-filled here from the original doc \u2014 its note flags that it looked strong on paper, so worth a second look before ruling it out for good.", False, 10),
    ("", False, 10),
    ("SUGGESTED WORKFLOW PER STUDIO", True, 12),
    ("1. Confirm contact / posting is still live.", False, 10),
    ("2. Pick portfolio pieces to feature (use the fit notes in \u2018Why It Fits\u2019 as a guide).", False, 10),
    ("3. Tailor CV + short intro email or website application.", False, 10),
    ("4. Send \u2014 set Status to \u2018Applied\u2019, fill Date Applied, set Follow-up Date (+2 weeks).", False, 10),
    ("5. On response, update Status (Responded / Interview / Offer / Rejected) and log it.", False, 10),
    ("6. Check the Dashboard's \u2018Follow-ups Due\u2019 count when you open the file \u2014 that's your action list.", False, 10),
    ("", False, 10),
    ("Some Location, Team Size and Contact cells are blank where the original notes didn't specify \u2014 confirm before applying rather than assuming.", False, 10),
]
for i, (text, bold, size) in enumerate(lines, start=1):
    cell = ws3.cell(row=i, column=1, value=text)
    cell.font = Font(name=FONT_NAME, bold=bold, size=size)
    cell.alignment = Alignment(wrap_text=True, vertical="top")

wb.active = 0
out_path = r"D:\JEAME SPACE\Internship Applications\Application Tracker.xlsx"
wb.save(out_path)
print("Saved:", out_path)
